#!/usr/bin/env python3
"""Simple updater for 'Диаграмма Ганта (1).xlsx'.
This script searches for a sheet named 'Диаграмма Ганта' (or first sheet),
looks for columns that resemble an ID column and a percent column and updates
ID 43101 from 0.65 to 0.78 (example). Run with Python and openpyxl installed.
"""
import sys
from pathlib import Path
try:
    import openpyxl
except Exception as e:
    print('openpyxl required: pip install openpyxl')
    raise

FILE = Path('Диаграмма Ганта (1).xlsx')
if not FILE.exists():
    FILE = Path(__file__).parent.parent / 'Диаграмма Ганта (1).xlsx'

if not FILE.exists():
    print('Excel file not found:', FILE)
    sys.exit(1)

wb = openpyxl.load_workbook(FILE)
sheet = None
if 'Диаграмма Ганта' in wb.sheetnames:
    sheet = wb['Диаграмма Ганта']
else:
    sheet = wb[wb.sheetnames[0]]

headers = {}
for col in sheet.iter_cols(min_row=1, max_row=1):
    val = (col[0].value or '').strip() if col[0].value else ''
    headers[val.lower()] = col[0].column

def find_column_by_keywords(keys):
    for h, col in headers.items():
        for k in keys:
            if k in h:
                return col
    return None

id_col = find_column_by_keywords(['номер','id','иср','номер иср'])
pct_col = find_column_by_keywords(['процент','%','выполн','percent'])

if not id_col or not pct_col:
    id_col = id_col or 1
    pct_col = pct_col or 4

target_id = '43101'
updated = False
for row in sheet.iter_rows(min_row=2):
    cell_id = str(row[id_col-1].value).strip() if row[id_col-1].value is not None else ''
    if cell_id == target_id:
        row[pct_col-1].value = 0.78
        updated = True
        break

if updated:
    wb.save(FILE)
    print(f'Updated {target_id} to 78% in {FILE}')
else:
    print('ID not found or not updated')
