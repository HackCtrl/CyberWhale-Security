#!/usr/bin/env python3
"""Create or open a local task from Gantt ID.
Usage: py -3 scripts/automation/take_task.py 43101
"""
from pathlib import Path
import sys, json
try:
    import openpyxl
except Exception:
    pass
ROOT = Path(__file__).resolve().parents[2]
GANTT = ROOT / 'Диаграмма Ганта (1).xlsx'
DB = ROOT / 'TASKS_DB.json'
TASKS_ACTIVE = ROOT / 'tasks' / 'active'

if len(sys.argv) < 2:
    print('Usage: take_task.py <GANTT_ID>')
    sys.exit(2)

gid = sys.argv[1]
if not DB.exists():
    print('TASKS_DB.json not found')
    sys.exit(2)

db = json.loads(DB.read_text(encoding='utf-8'))
local_map = db.setdefault('local_mapping', {})
if gid in local_map:
    lid = local_map[gid]
    print('Task already mapped:', lid)
    sys.exit(0)

# Try to find row in Excel by gantt id or by name
if not GANTT.exists():
    print('Gantt file not found; will create minimal task')
    name = input('Введите название задачи: ')
else:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(GANTT, data_only=True)
        sheet = wb['Диаграмма Ганта'] if 'Диаграмма Ганта' in wb.sheetnames else wb.active
        header_row = 8
        found = None
        for row in sheet.iter_rows(min_row=header_row+1):
            try:
                cand = row[1].value
                if cand is None:
                    continue
                if str(cand).strip() == gid:
                    found = row
                    break
            except Exception:
                continue
        if found:
            name = found[2].value or f'Task {gid}'
        else:
            # try to use mapping from TASKS_DB.json -> gantt_mapping
            gantt_map = db.get('gantt_mapping', {})
            if gid in gantt_map:
                name = gantt_map[gid]
                print('Found name in TASKS_DB.gantt_mapping:', name)
            else:
                name = input('Gantt row not found; введите название задачи: ')
    except Exception:
        name = input('openpyxl not available; введите название задачи: ')

# allocate local id
last = db.get('last_id', 0) + 1
lid = f"{last:02d}"
local_map[gid] = lid
# create task file
TASKS_ACTIVE.mkdir(parents=True, exist_ok=True)
filename = TASKS_ACTIVE / f"{lid}-{''.join(c if c.isalnum() else '-' for c in str(name).lower())[:80]}.md"
content = []
content.append(f"# ТЗ-{lid}: {name}")
content.append('')
content.append(f"**Gantt ID**: {gid}")
content.append(f"**Локальный ID**: {lid}")
content.append('**Статус**: in_progress')
content.append('')
content.append('### Архитектура')
content.append('')
content.append('### API / DB изменения')
content.append('')
content.append('### Тесты (не менее 80% покрытия)')
content.append('')
content.append('### Критерии приёмки')
filename.write_text('\n'.join(content), encoding='utf-8')
# update db
db['last_id'] = last
db.setdefault('tasks', []).append({'local_id':lid,'gantt_id':gid,'name':name,'status':'in_progress','percent':0})
DB.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding='utf-8')
print('Created task', filename.name)
