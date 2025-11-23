#!/usr/bin/env python3
"""
Скрипт синхронизации локальной БД задач с Excel-файлом "Диаграмма Ганта (1).xlsx".

Как работает (по умолчанию):
- открывает книгу `Диаграмма Ганта (1).xlsx` (в корне проекта);
- читает первую таблицу; строки — задачи;
- извлекает Gantt ID, заголовок и процент выполнения (параметры колонок настраиваемы);
- для новых задач создаёт local ID (двузначный), файл в `tasks/active` на основе `tasks/TASK_TEMPLATE.md`;
- обновляет `TASKS_DB.json` и `ROADMAP.md`.

Примечание: при других вариантах расположения колонок — отредактируйте константы ниже.
"""
import json
from pathlib import Path
from datetime import datetime
import re
import argparse

try:
    import openpyxl
except Exception:
    print("Ошибка: требуется пакет openpyxl. Установите: python -m pip install openpyxl")
    raise

ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL = ROOT / "Диаграмма Ганта (1).xlsx"
DB_FILE = ROOT / "TASKS_DB.json"
TASKS_DIR = ROOT / "tasks" / "active"
TEMPLATE_FILE = ROOT / "tasks" / "TASK_TEMPLATE.md"
ROADMAP_FILE = ROOT / "ROADMAP.md"

# Настройки: подставьте индексы колонок (1-based)
GANTT_ID_COL = 1
NAME_COL = 2
PERCENT_COL = 3
HEADER_ROW = 2


def load_db():
    if not DB_FILE.exists():
        return {"schema_version": 1, "last_id": 0, "tasks": [], "local_mapping": {}}
    return json.loads(DB_FILE.read_text(encoding="utf-8"))


def save_db(db):
    DB_FILE.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")


def next_local_id(db):
    db["last_id"] = db.get("last_id", 0) + 1
    return f"{db['last_id']:02d}"


def slugify(s: str):
    s = s.lower()
    s = re.sub(r"[^a-z0-9а-яё]+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def create_task_file(local_id, gantt_id, title, percent):
    TASKS_DIR.mkdir(parents=True, exist_ok=True)
    slug = slugify(title)[:40]
    filename = TASKS_DIR / f"{local_id}-{slug}.md"
    created = datetime.utcnow().isoformat()
    if TEMPLATE_FILE.exists():
        tmpl = TEMPLATE_FILE.read_text(encoding="utf-8")
        content = tmpl.format(title=title, local_id=local_id, gantt_id=gantt_id or "", percent=percent or 0, created=created)
    else:
        content = f"---\ntitle: \"{title}\"\nlocal_id: {local_id}\ngantt_id: {gantt_id}\npercent: {percent}\ncreated: {created}\n---\n\n# {title}\n"
    filename.write_text(content, encoding="utf-8")
    return filename


def regenerate_roadmap(db):
    lines = []
    lines.append("# ROADMAP — Живая карта задач\n")
    lines.append("| Local ID | Gantt ID | Заголовок | Статус | % | Исполнитель |")
    lines.append("|----------|----------|-----------|--------|---|-------------|")
    for t in db.get("tasks", []):
        lines.append(f"| {t.get('local_id','')} | {t.get('gantt_id','')} | {t.get('title','')} | {t.get('status','')} | {t.get('percent',0)} | {t.get('assignee','')} |")
    ROADMAP_FILE.write_text("\n".join(lines), encoding="utf-8")


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=str(EXCEL), help="Путь к файлу диаграммы Ганта")
    parser.add_argument("--sheet", default=None, help="Имя листа (по умолчанию первый)")
    args = parser.parse_args(argv)

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    sheet = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    db = load_db()
    created = 0
    updated = 0

    for r in range(HEADER_ROW, sheet.max_row + 1):
        gantt_id = sheet.cell(row=r, column=GANTT_ID_COL).value
        title = sheet.cell(row=r, column=NAME_COL).value
        percent = sheet.cell(row=r, column=PERCENT_COL).value
        if title is None:
            continue
        gid = str(gantt_id).strip() if gantt_id is not None else ""
        # normalize percent
        try:
            p = float(percent) if percent is not None else 0.0
            if p > 1:
                p = round(p, 2)
            else:
                p = round(p * 100, 2)
        except Exception:
            p = 0.0

        # existing mapping?
        local = db.get("local_mapping", {}).get(gid) if gid else None
        if local:
            # update task percent in DB
            for t in db.get("tasks", []):
                if t.get("local_id") == local:
                    t["percent"] = p
                    updated += 1
                    break
            continue

        # new task
        local_id = next_local_id(db)
        task = {"local_id": local_id, "gantt_id": gid, "title": str(title), "status": "todo", "percent": p, "assignee": ""}
        db.setdefault("tasks", []).append(task)
        if gid:
            db.setdefault("local_mapping", {})[gid] = local_id
        create_task_file(local_id, gid, str(title), p)
        created += 1

    save_db(db)
    regenerate_roadmap(db)
    print(f"Sync complete. created={created}, updated={updated}, last_id={db.get('last_id')}")


if __name__ == '__main__':
    main()
